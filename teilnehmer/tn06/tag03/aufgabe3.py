#!/usr/bin/env python3

import csv
import openpyxl
import pprint
#from openpyxl.cell import get_column_letter

dateiname = "HistoricalQuotes.csv"
zieldatei = "Excel.xlsx"
try:
    zeilen = []
    with open("HistoricalQuotes.csv", newline='') as f:
        daten = csv.reader(f, delimiter=",")
        for zeile in daten:
            #print(zeile)
            zeilen.append(zeile)
    for pos, zeile in enumerate(zeilen):
        #print(zeile)
        if pos > 0:
            datum_alt=zeile[0].split('/')
            zeile[0] = f"{datum_alt[1]}.{datum_alt[0]}.{datum_alt[2]}"
            zeile[1] = float(zeile[1].replace("$",""))
            zeile[2] = int(zeile[2])
            zeile[3] = float(zeile[3].replace("$",""))
            zeile[4] = float(zeile[4].replace("$",""))
            zeile[5] = float(zeile[5].replace("$",""))

        print(zeile)

    
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "Apple-Aktie"
    for row_index, zeile in enumerate(zeilen):
        for column_index, zelle in enumerate(zeile):
            ws.cell(column=column_index+1, row=row_index+1, value=zelle)
    wb.save(filename = zieldatei)

except Exception as e:
    print (e)
