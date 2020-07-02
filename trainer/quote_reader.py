#!/usr/bin/env python3
import csv
import re
import openpyxl

fname = "/home/coder/python_grundlagen/materialien/HistoricalQuotes.csv"

def read_quotes(fname):
    ergebnis = []

    with open(fname, newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            ergebnis.append(row)
    return ergebnis

def verarbeitung(daten):
    ergebnis = []
    for row in daten:
        d = dict(row.copy())
        for key in d.keys():
            value = d[key].strip()
            m = re.search(r"^(\d{2})/(\d{2})/(\d{4})$", value)
            if m:
                d[key] = f"{m.group(2)}.{m.group(1)}.{m.group(3)}"
            elif "$" in value:
                d[key] = float(value.replace("$", ""))
            elif value.isnumeric():
                d[key] = int(value)
        ergebnis.append(d)
    return ergebnis

def build_xls(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apple"
    # Überschriften
    for col, name in enumerate(data[0].keys(), 1):
        ws.cell(column=col, row=1, value=name)
    # Daten
    for row, row_values in enumerate(data, 2):
        for col, value in enumerate(row_values.values(), 1):
            ws.cell(column=col, row=row, value=value)
    return wb

def write_xls(fname, wb):
    wb.save(fname)
    
if __name__ == "__main__":
    quotes = read_quotes(fname)
    quotes = verarbeitung(quotes)
    xls = build_xls(quotes)
    write_xls("apple.xlsx", xls)
